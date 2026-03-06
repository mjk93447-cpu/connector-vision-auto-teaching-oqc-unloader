"""
Excel format inference and output.
Read training Excel, infer structure, write inference results in same format.
Multi-row format: one Excel file, row 1=headers, row 2+ = one per cell (A2HDxxxx).
"""
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
except ImportError:
    openpyxl = None

# Cell ID column names (Korean/English)
CELL_COLUMN_NAMES = ("셀", "cell", "cellid", "cell_id", "a2hd", "번호", "no", "num")


def load_excel_format(excel_path: str | Path) -> Dict[str, Any]:
    """
    Load Excel and infer structure (headers, column indices).
    Returns dict with: headers, sheet_name, sample_row.
    """
    if openpyxl is None:
        raise ImportError("openpyxl required: pip install openpyxl")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    sample = [c.value for c in ws[2]] if ws.max_row >= 2 else []
    wb.close()
    return {
        "headers": headers,
        "sheet_name": ws.title,
        "sample_row": sample,
    }


def find_cell_column_index(headers: List[str]) -> int | None:
    """Find column index for cell ID (A2HDxxxx). Returns None if not found."""
    for i, h in enumerate(headers):
        h_lower = str(h or "").lower().replace(" ", "").replace("_", "")
        for name in CELL_COLUMN_NAMES:
            if name in h_lower or (name == "a2hd" and "a2hd" in h_lower):
                return i
    return None


def infer_column_indices(headers: List[str]) -> Dict[str, int]:
    """
    Infer column indices from headers.
    Korean: 위핀, 아래핀, 위핀개수, 아래핀개수, OK, NG, 판정, 좌우간격, spacing.
    English: upper, lower, upper_count, lower_count, judgment, spacing.
    """
    mapping = {}
    for i, h in enumerate(headers):
        h_lower = h.lower().replace(" ", "").replace("_", "")
        # Korean
        if "위" in h and ("핀" in h or "개" in h):
            mapping["upper_count"] = i
        elif "아래" in h and ("핀" in h or "개" in h):
            mapping["lower_count"] = i
        elif "ok" in h_lower or "ng" in h_lower or "판정" in h:
            mapping["judgment"] = i
        elif "간격" in h or "spacing" in h_lower or "거리" in h:
            mapping["spacing"] = i
        # English
        if "upper" in h_lower and ("count" in h_lower or "pin" in h_lower or "num" in h_lower):
            mapping["upper_count"] = i
        elif "lower" in h_lower and ("count" in h_lower or "pin" in h_lower or "num" in h_lower):
            mapping["lower_count"] = i
        elif "judgment" in h_lower or "result" in h_lower or "verdict" in h_lower:
            mapping["judgment"] = i
    return mapping


def load_excel_multi_row(excel_path: str | Path) -> Dict[str, Any]:
    """
    Load Excel with multiple data rows (one per cell).
    Returns: headers, rows (list of row values), cell_col_idx, sheet.
    """
    if openpyxl is None:
        raise ImportError("openpyxl required: pip install openpyxl")
    wb = load_workbook(excel_path, read_only=False, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append([ws.cell(row=r, column=c + 1).value for c in range(len(headers))])
    cell_col = find_cell_column_index(headers)
    return {
        "headers": headers,
        "rows": rows,
        "cell_col_idx": cell_col,
        "workbook": wb,
        "sheet": ws,
    }


def find_row_index_by_cell_id(rows: List[List], cell_id: str, cell_col_idx: int | None) -> int | None:
    """Find row index where cell column matches cell_id. 0-based."""
    if cell_col_idx is None:
        return None
    cell_upper = str(cell_id or "").upper()
    for i, row in enumerate(rows):
        if cell_col_idx < len(row):
            val = str(row[cell_col_idx] or "").strip().upper()
            if val == cell_upper or cell_upper in val:
                return i
    return None


def write_result_excel(
    output_path: str | Path,
    upper_count: int,
    lower_count: int,
    upper_spacings: List[float],
    lower_spacings: List[float],
    format_ref: Optional[Dict[str, Any]] = None,
    cell_id: Optional[str] = None,
    update_existing: bool = False,
) -> None:
    """
    Write inference result to Excel.
    If format_ref from load_excel_format, use same structure; else minimal format.
    Multi-row: if update_existing and cell_id, load existing Excel and update row for that cell.
    """
    if openpyxl is None:
        raise ImportError("openpyxl required: pip install openpyxl")
    judgment = "OK" if (upper_count == 20 and lower_count == 20) else "NG"
    spacing_str = ", ".join(f"{s:.2f}" for s in (upper_spacings + lower_spacings))
    upper_str = ", ".join(f"{s:.2f}" for s in upper_spacings)
    lower_str = ", ".join(f"{s:.2f}" for s in lower_spacings)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if update_existing and cell_id and output_path.exists():
        data = load_excel_multi_row(output_path)
        idx = infer_column_indices(data["headers"])
        wb = data.get("workbook")
        ws = data["sheet"]
        row_idx = find_row_index_by_cell_id(data["rows"], cell_id, data.get("cell_col_idx"))
        if wb and ws:
            if row_idx is not None:
                r = row_idx + 2  # 1-based, skip header
                if "upper_count" in idx:
                    ws.cell(row=r, column=idx["upper_count"] + 1, value=upper_count)
                if "lower_count" in idx:
                    ws.cell(row=r, column=idx["lower_count"] + 1, value=lower_count)
                if "judgment" in idx:
                    ws.cell(row=r, column=idx["judgment"] + 1, value=judgment)
                if "spacing" in idx:
                    ws.cell(row=r, column=idx["spacing"] + 1, value=spacing_str)
            else:
                # Append new row for this cell
                cell_col = data.get("cell_col_idx")
                new_row = [""] * len(data["headers"])
                if cell_col is not None:
                    new_row[cell_col] = cell_id
                if "upper_count" in idx:
                    new_row[idx["upper_count"]] = upper_count
                if "lower_count" in idx:
                    new_row[idx["lower_count"]] = lower_count
                if "judgment" in idx:
                    new_row[idx["judgment"]] = judgment
                if "spacing" in idx:
                    new_row[idx["spacing"]] = spacing_str
                ws.append(new_row)
            wb.save(output_path)
            return

    wb = Workbook()
    ws = wb.active
    if format_ref and format_ref.get("headers"):
        headers = format_ref["headers"]
        ws.append(headers)
        row = [""] * len(headers)
        idx = infer_column_indices(headers)
        if cell_id and (cell_col := find_cell_column_index(headers)) is not None:
            row[cell_col] = cell_id
        if "upper_count" in idx:
            row[idx["upper_count"]] = upper_count
        if "lower_count" in idx:
            row[idx["lower_count"]] = lower_count
        if "judgment" in idx:
            row[idx["judgment"]] = judgment
        if "spacing" in idx:
            row[idx["spacing"]] = spacing_str
        ws.append(row)
    else:
        headers = ["셀", "위핀개수", "아래핀개수", "판정", "위핀간격(mm)", "아래핀간격(mm)"]
        ws.append(headers)
        row = [cell_id or "", upper_count, lower_count, judgment, upper_str, lower_str]
        ws.append(row)
    wb.save(output_path)
